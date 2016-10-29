from decimal import *
from openpyxl import Workbook, load_workbook

from .models import *
from .logger import *

def name2Id ( rname, rtype ):
	""" для суперсправочников: находит ID по имени элемента, 
		если элемент не существует - создает его
	"""

	for r in Refs.objects.filter( rf_type=rtype, rf_descr=rname ):
		return r.id

	ref = Refs.objects.create(
		rf_type = rtype,
		rf_descr = rname
	)
	return ref.id

def	addAttrVal ( iid, aname, grpid, aval ):
	""" добавляет значение атрибута для товара, создает атрибут, если он не существует """

	at_sname = models.CharField('атрибут',max_length=100)
	at_grp = models.ForeignKey(Refs,verbose_name='группа товаров')
	at_descr = models.CharField('описание',max_length=10000,blank=True, null=True)

	aid = None
	for a in Attrs.objects.filter( at_grp=grpid, at_sname=aname ):	# может ли быть несколько???
		aid = a.id

	if aid is None:		# новый атрибут - создаем
		a = Attrs.objects.create(
			at_grp_id = grpid,
			at_sname = aname,
			at_descr = "Создано в процессе импорта"
		)
		aid = a.id

	# создаем значение атрибута
	av = AtVals.objects.create(
		av_item_id = iid,
		av_attr_id = aid,
		av_value = aval
	)

def getItems(fn, sh):
	""" создает товары из файла и его закладки
	формат файла:
		- сначала идут постоянные колонки (смысл которых фиксирован, см. ниже), к примеру сейчас это
			код
			ориг код
			производитель (ID)
			краткое наименование
			цена
			категория
			группа
		- потом - переменные с атрибутами (разные наборы атрибутов для разных товаров)
			переменные идут в виде имя атрибута в заголовке, значение в этой колонке
			заголовок переменных колонок всегда начинается с решетки (т.е. это начало списка названий атрибутов)
	возвращает текст ошибки или пусто (если ок)
	"""

	ncAttr = 8
	# открываем файл с данными
	wb = load_workbook(fn)
	try:
		ws = wb.get_sheet_by_name(sh)
	except:
		return "В файле нет закладки с именем '{0}'".format(sh)

	# проходим по строкам файла, создаем товары
	for l in range(1,10000):	# КОСТЫЛЬ: пока ограничение - надо будет заменить на макс строку файла, такое есть...
		if ws.cell(row=l,column=ncAttr).value is None:	# даные кончились
			break

		if str(ws.cell(row=l,column=ncAttr).value)[0]=='#':	 # строка содержит названия атрибутов - заполняем словарь
			atrs = []
			nAttr = 0
			for c in range(ncAttr,100):
				if ws.cell(row=l,column=c).value is None:	# даные кончились
					break
				atrs.append(str(ws.cell(row=l,column=c).value).replace('#',''))		# запоминаем имя атрибута - удаляем решетки
				nAttr += 1
		else:	# строка содержит данные
			# создаем товар
			# DoLog("#5="+str(ws.cell(row=l,column=5).value))
			item = Items.objects.create(
				it_code = str(ws.cell(row=l,column=1).value),
				it_orig_code = str(ws.cell(row=l,column=2).value),
				it_man_id = ws.cell(row=l,column=3).value,
				it_sname = str(ws.cell(row=l,column=4).value),
				it_base_price = Decimal(str(ws.cell(row=l,column=5).value)),
				it_cat_id = name2Id ( str(ws.cell(row=l,column=6).value), RefType.CAT.value ),
				it_grp_id = name2Id ( str(ws.cell(row=l,column=7).value), RefType.GRP.value )
			)

			# создаем атрибуты и их значения (добавляя атрибуты, если таковые не существуют)
			for c in range(ncAttr,ncAttr+nAttr):
				addAttrVal ( item.id, atrs[c-ncAttr], item.it_grp_id, str(ws.cell(row=l,column=c).value) )

	return ""	# признак ОК


